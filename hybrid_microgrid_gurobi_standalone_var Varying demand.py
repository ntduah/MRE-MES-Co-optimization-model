'''Created on Mar 30, 2024
#RE is varied along wuth the stochastic demand
@author: yangc
'''
import openpyxl
import xlsxwriter
import numpy as np
import os
import math
import time
import random
random.seed(10)
import matplotlib.pyplot as plt
import gurobipy as gp
from gurobipy import GRB


global Exp_num
Exp_num = 1000 
Random_range = 0.05                                                                                                                                                                                


class Parameter():
    def __init__(self):
        """ define parameters and read data """
        self.Time_period = 336      # time period
        self.days = 14
        self.DeltaT = 1             # time step 1 hour
        self.BigM = 1000           # 'big number for MIP'               /1000/ 
        self.SmallN = 0.0000001    # 'small number for output'         /0.0000001/ 
        self.eta_WT = 0.35         # 'efficiency oof wind turbine'    /0.35/ 
        self.rho_WT = 1.17682      # 'parameter of wind power generation, g/m3' /1.17682/     
        self.PI = 3.1415           # 'number pi'                     /3.1415/ 
        self.DWT = 58.13           #'diameter of wind turbine,m'     /58.13/ 

        self.Cbmax = 1            #'capacity of battery storage'  /1/
        self.Cqmax = 1.5            #'capacity of thermal storage'  /1.5/
        self.Cwmax = 50            #'capacity of water tank'       /50/
        self.Chmax = 200            #'capacity of hydrogen tank'    /200/
        
        self.carbon_price = 0.18      # 60 euro/ton => 0.072 $/kg  #https://www.imf.org/en/Publications/WP/Issues/2024/02/09/Carbon-Prices-and-Inflation-in-the-Euro-Area-544465
        self.gas_emmi_factor = 1.9423  # 0.055kg/cubic foot => 1.9423 kg/m3
        self.gas_price = 0.64       # range (9-18), $9/thousand cubic feet => $0.3178/m3#https://www.eia.gov/dnav/ng/hist/n3020nc3m.htm
        
        self.life_time = 50         # life time 50 years
        self.interest_rate = 0.05       # interest rate
       
        '''parameters for wind turbine '''
        self.I_w = 3                # cut-in velocity I for wind turbine  m/s
        self.Phi_w = 25             # cut-off velocity Phi for wind turbine  m/s
        self.Lambda_w = 13          # rated velocity lambda for wind turbine  m/s
        self.Ucost_w = 1300000      # unit cost of wind turbine  $/mw test: 3270000, #https://weatherguardwind.com/how-much-does-wind-turbine-cost-worth-it/#:~:text=%241%2C300%2C000%20USD%20per%20megawatt.,%242%2D4%20million%20dollar%20range.
        self.Rated_w = 0.5          # rated power of one wind turbine MW, test: 0.5 original value 2.5 #https://weatherguardwind.com/how-much-does-wind-turbine-cost-worth-it/#:~:text=%241%2C300%2C000%20USD%20per%20megawatt.,%242%2D4%20million%20dollar%20range.
        self.Mcost_w = 20000        # maintance cost of wind turbine $/MW/year 
        
        '''parameters for solar panel'''
        self.tal_PV = 0.9          #'para of pv panel'             /0.9/
        self.eta_PV = 0.15         # 'effi of pv panel'             /0.15/
        self.gamma_PV = 0.0045        # 'para of pv panel'             /0.0045/ 
        self.cap_pv_m2 = 0.0002       # power capacity in MW for 1 m2 solar panel #https://www.architecturaldigest.com/reviews/solar/how-much-power-does-a-solar-panel-produce#:~:text=Solar%20panels%20are%20rated%20by%20the%20amount%20of,produce%20200%20to%20300%20watts%20per%20square%20meter.
        self.Ucost_PV = 30         # unit cost of pv panel        $/m2
        self.Mcost_PV = 13200         # maintance cost of pv panel $/MW/year
        
        '''parameters for wave energy converter'''
        self.water_depth = 30     # water depth, shallow water up to 50m
        self.gravity_accel = 9.8  # gravitational acceleration m/s2
        self.water_density = 1028  # seawater mass density, kg/m3
        self.eta_wave = 0.4        #overall efficiency of wave converter
        self.water_front = 120     # overall wave contact length
        self.Ucost_wave = 3000000    # $/MW 3000000 #https://wacop.gsd.spc.int/RegionalWaveEnergy.html
        self.wave_time = 2       # second
        self.Mcost_wave = 0.016    # maintance cost of wave energy converter % of capital investment #https://wacop.gsd.spc.int/RegionalWaveEnergy.html

        '''parameters for water desalination'''
        self.C0 = 36000               #'initial salt concentration of feed water, ppm' /36000/
        self.eta_des = 0.55          #'water recovery ratio' /0.55/
        self.mem_resist = 33.95       # membrane resistance Pa.sec/m
        self.Mcost_des = 0.02         # annual maintance cost of desalination % of capital investment #chno-economic assessment of electrodialysis and reverse osmosis desalination plants
        
        '''parameters for CHP system'''
        self.eta_P = 0.35           #'effi of gas-power'          /0.35/
        self.eta_Q = 0.65           #'effi of gas-power'          /0.65/
        self.eta_L = 0.5           #'effi of gas-power'          /0.5/
        self.eta_X = 0.5           #'effi of gas-power'          /0.5/        
        self.Vgas = 0.0097             #'low calorific value of gas, 9.7kwh/m3' /0.0097/        
        # self.Mcost_boiler = 2.7            # maintance cost of boiler $/MW 
        self.Mcost_boiler = 0.432 * 1000 * 0.145 # $/MW, see table 2-3 in paper "see paper "Study on the optimization and sensitivity analysis of CCHP systems for T industrial park facilities""
        # self.Mcost_turbine = 20000    # maintance cost of turbine $/MW https://www.power-technology.com/features/featurepower-plant-om-how-does-the-industry-stack-up-on-cost-4417756/?cf-view
        self.Mcost_turbine = 0.03     # see table 2-3 in paper "Study on the optimization and sensitivity analysis of CCHP systems for T industrial park facilities"
        
        '''parameters for electrolyzer'''
        self.eta_eltz = 0.60        # 'effi of electrolyzer'       /0.60/
        self.PH =  360              #'power to hydrogen conversion factor, m3/mwh' /360/
        self.alpha_lb = 0         # lower bound coefficient for power consumption of electrolyzer
        self.rated_electzer = 1.2   # rated power capacity MW
        self.ce_electzer = 1500000  # specific cost of electrolyzer $/mw        
        self.Mcost_electzer = 0.02  # annual maintance cost of electzer % of capital investment $Assessment of Hydrogen Production Costs from Electrolysis: United States and Europe
    
        '''parameters for storages and tanks'''
        
        self.Ucost_Swater = 250   # $/m3, #https://thundersaidenergy.com/downloads/storage-tank-costs-storing-oil-energy-water-and-chemicals/
        self.Ucost_Shydro = 103.75 # 1250*0.083$/m3 = (1250$/kg) * (70.85kg/m3 in liquid density, and 0.083 kg/m3 in gas)  #Techno-economic assessment of solar PVfuel cell hybrid system for telecom base stations in Ghana.pdf 
        # based on market: https://www.alibaba.com/product-detail/cryogenic-storage-tank-Cryogenic-liquid-oxygen_1600470309328.html
        self.Ucost_Stherm = 622000    # $/mw  check table 3 of  'Levelised Cost of Storage for Pumped Heat Energy Storage in comparison with other energy storage technologies   
        #https://www.energy.gov/eere/solar/project-profile-innovative-thermal-energy-storage-baseload-solar-power-generation
        self.Ucost_Sbatt = 1000000   #$/Mw  #https://howtostoreelectricity.com/costs-of-1-mw-battery/
        # https://thundersaidenergy.com/2023/11/18/grid-scale-battery-costs-kw-or-kwh/
        # https://www.nrel.gov/docs/fy21osti/79236.pdf
        
        self.Mcost_Swater = 0.01  #maintance cost of water tank % of capital investment, see #Technical-economic framework for designing of water pumping system based on photovoltaic clean energy with water storage for drinking application
        self.Mcost_Shydro = 0.8466  #maintance cost of hydrogen tank 10.2 $/kg/year * (0.083 kg/m3 in gas) = 0.8466$/m3 in gas 
        # A novel hybrid optimization framework for sizing renewable energy systems integrated with energy storage systems with solar photovoltaics, wind, battery and electrolyzer-fuel cell
        # Techno-economic assessment of solar PVfuel cell hybrid system for telecom base stations in Ghana.pdf 
        self.Mcost_Stherm = 0.01  #maintance cost of thermal storage % of capital investment #Evidence Gathering:Thermal Energy Storage (TES) Technologies
        self.Mcost_Sbatt = 8000  #maintance cost of batt storage $/Mw/year #https://www.pnnl.gov/sites/default/files/media/file/Final%20-%20ESGC%20Cost%20Performance%20Report%2012-11-2020.pdf

        self.cdr = 0.25              #'charging, discharging rate' /0.25/
        self.eta_bd = 0.9           #'discharge efficiency' /0.9/
        self.eta_bc = 0.9           #'charge efficiency' /0.9/
        self.eta_qd = 0.85           #'discharge efficiency' /0.85/
        self.eta_qc = 0.85           #'charge efficiency' /0.85/
        self.eta_wd = 1           #'discharge efficiency' /1/
        self.eta_wc = 1           #'charge efficiency' /1/
        self.eta_hd = 1           #'discharge efficiency' /1/
        self.eta_hc = 1           #'charge efficiency' /1/
        self.IB = 0.2               #'initial battery level'
        self.IQ = 0.2               #'initial thermal level'
        self.IW = 0.2               #'initial water level'
        self.IH = 0.2                #'initial hydrogen level
        
        '''parameters for Purchasing hydrogen, water, thermal'''
        self.Lambda_Q = 104        #purchasing cost $/MWh of thermal energy through electricity #https://www.deltat.com/thermal_energy_costs.html
        self.Lambda_H = 1.37033         # purchasing cost 16.51 $/kg * (0.083 kg/m3 in gas) = 1.37033 $/m3  #https://h2fcp.org/category/hydrogen-stations
        self.Lambda_W = 1.062     #Purchasing cost $/m3   #https://thecameronteam.net/how-much-is-the-average-water-bill-in-wilmington-nc/
        
        self.end_day_time_list = [(i*24-1) for i in range(1,self.days+1)]
        
        '''parameters for degradation of energy storage'''
        self.gamma_B = 8        # $/MW
        self.gamma_H = 0.05     # $/m3
        self.gamma_Q = 2         # $/MW
        self.gamma_W = 0.05      # $/m3
        
        'Electricity price'
        self.Elec_price = {}
        f_name = 'electricity_price_2weeks'
        # Load the workbook and select the active sheet
        book = openpyxl.load_workbook(f_name + '.xlsx')
        # Select the second sheet by index
        sheet = book.worksheets[0]
        # Read data from the sheet
        for t in range(self.Time_period):
            self.Elec_price[t] = sheet.cell(row=t+2, column=2).value  # Adjust indexing for 1-based indexing in openpyxl
        # No need to release resources explicitly in openpyxl    
        
        'wave height'
        self.HWave = {}
        f_name = 'waveheight_2weeks'
        book = openpyxl.load_workbook(f_name + '.xlsx')
        sheet = book.worksheets[0]
        for t in range(self.Time_period):
            self.HWave[t] = sheet.cell(row=t+2, column=2).value * 1.5 
            self.HWave[t] = random.uniform(self.HWave[t] * (1 - 0.05), self.HWave[t] * (1 + 0.05)) 
    
        'wind speed'
        self.VWind = {}
        f_name = 'windspeed_2weeks'
        book = openpyxl.load_workbook(f_name + '.xlsx')
        sheet = book.worksheets[0]
        for t in range(self.Time_period):
            self.VWind[t] = sheet.cell(row=t+2, column=2).value  * 1.5
            self.VWind[t] = random.uniform(self.VWind[t] * (1 - 0.05), self.VWind[t] * (1 + 0.05))        

        'solar power'
        self.solar_irri = {}
        f_name = 'solarirridiance_2weeks'
        book = openpyxl.load_workbook(f_name + '.xlsx')
        sheet = book.worksheets[0]
        for t in range(self.Time_period):
            self.solar_irri[t] = sheet.cell(row=t+2, column=2).value  
            self.solar_irri[t] = random.uniform(self.solar_irri[t] * (1 - 0.05), self.solar_irri[t] * (1 + 0.05))  

        'temperature'
        self.temp = {}
        f_name = 'temperature_2weeks'
        book = openpyxl.load_workbook(f_name + '.xlsx')
        sheet = book.worksheets[0]
        for t in range(self.Time_period):
            self.temp[t] = sheet.cell(row=t+2, column=2).value  
            self.temp[t] = random.uniform(self.temp[t] * (1 - 0.05), self.temp[t] * (1 + 0.05))       
        
        'thermal demand'
        self.qD = {}
        f_name = 'thermal_demand_2weeks'
        book = openpyxl.load_workbook(f_name + '.xlsx')
        sheet = book.worksheets[0]
        for t in range(self.Time_period):
            self.qD[t] = sheet.cell(row=t+2, column=2).value
            self.qD[t] = random.uniform(self.qD[t] * (1 - Random_range), self.qD[t] * (1 + Random_range))           

        'hydrogen demand'
        self.hD = {}
        f_name = 'hydrogen_demand_2weeks'
        book = openpyxl.load_workbook(f_name + '.xlsx')
        sheet = book.worksheets[0]
        for t in range(self.Time_period):
            self.hD[t] = sheet.cell(row=t+2, column=2).value  
            self.hD[t] = random.uniform(self.hD[t] * (1 - Random_range), self.hD[t] * (1 + Random_range))   

        'power demand'
        self.eD = {}
        f_name = 'power_demand_2weeks'
        book = openpyxl.load_workbook(f_name + '.xlsx')
        sheet = book.worksheets[0]
        for t in range(self.Time_period):
            self.eD[t] = sheet.cell(row=t+2, column=2).value  
            self.eD[t] = random.uniform(self.eD[t] * (1 - Random_range), self.eD[t] * (1 + Random_range))   
        
        'water demand'
        self.wD = {}
        f_name = 'water_demand_2weeks'
        book = openpyxl.load_workbook(f_name + '.xlsx')
        sheet = book.worksheets[0]
        for t in range(self.Time_period):
            self.wD[t] = sheet.cell(row=t+2, column=2).value 
            self.wD[t] = random.uniform(self.wD[t] * (1 - Random_range), self.wD[t] * (1 + Random_range))   
        
        # print(self.wD)
        # print(max(self.wD))
        
        # self.P_Sbatt = max(self.eD.values())  *0.000001
        # self.P_Stherm = max(self.qD.values()) *0.000001 
        # self.V_Swater = max(self.wD.values()) *0.000001
        # self.V_Shydro = max(self.hD.values()) *0.000001
        
        # self.P_Sbatt = 0.000001
        # self.P_Stherm = 0.0000010
        # self.V_Swater = 0.000001
        # self.V_Shydro = 0.000001
        
        # print(self.P_Sbatt)
        # print(self.P_Stherm)
        # print(self.V_Swater)
        # print(self.V_Shydro)
        
        # self.Water_flow_bound = 1500/24 # see section "Case Study #4: ROX Plant 2" in paper "Operation modeling and comparison of actual multi-effect distillation and reverse osmosis desalination plants"
        self.Water_flow_bound = 11000/24 # see figure 7 in "Design and performance simulation comparison of a wave energy-powered and wind-powered modular desalination system""
     

if __name__ == '__main__':
    begin_time = time.time()
    workbook0 = xlsxwriter.Workbook('MES_result_gurobi_standalone__0%RE_unlimit_PV_hydro_water_vss.xlsx')
    worksheet0_0 = workbook0.add_worksheet('obj')
    worksheet0_1 = workbook0.add_worksheet('b_x')
    worksheet0_2 = workbook0.add_worksheet('q_x')
    worksheet0_3 = workbook0.add_worksheet('w_x')
    worksheet0_4 = workbook0.add_worksheet('h_x')
    worksheet0_5 = workbook0.add_worksheet('b_xd')
    worksheet0_6 = workbook0.add_worksheet('b_xc')
    worksheet0_7 = workbook0.add_worksheet('q_xd')
    worksheet0_8 = workbook0.add_worksheet('q_xc')
    worksheet0_9 = workbook0.add_worksheet('w_xd')
    worksheet0_10 = workbook0.add_worksheet('w_xc')
    worksheet0_11 = workbook0.add_worksheet('h_xd')
    worksheet0_12 = workbook0.add_worksheet('h_xc')   
    worksheet0_13 = workbook0.add_worksheet('rPow')
    worksheet0_14 = workbook0.add_worksheet('wFlow')
    worksheet0_15 = workbook0.add_worksheet('eP_des')
    worksheet0_16 = workbook0.add_worksheet('q1GT')
    worksheet0_17 = workbook0.add_worksheet('q2GT')
    worksheet0_18 = workbook0.add_worksheet('pGT')
    worksheet0_19 = workbook0.add_worksheet('gFlowP')
    worksheet0_20 = workbook0.add_worksheet('gFlowB')
    worksheet0_21 = workbook0.add_worksheet('hFlow')
    worksheet0_22 = workbook0.add_worksheet('eP_eltz')   
    worksheet0_23 = workbook0.add_worksheet('sB')
    worksheet0_24 = workbook0.add_worksheet('sQ')
    worksheet0_25 = workbook0.add_worksheet('sW')
    worksheet0_26 = workbook0.add_worksheet('sH')  
    
    worksheet0_27 = workbook0.add_worksheet('N_wind')
    worksheet0_28 = workbook0.add_worksheet('A_pv')
    worksheet0_29 = workbook0.add_worksheet('N_wave')
    worksheet0_30 = workbook0.add_worksheet('N_osmosis')
    worksheet0_31 = workbook0.add_worksheet('N_electzer')
    worksheet0_32 = workbook0.add_worksheet('p_turbine')
    worksheet0_33 = workbook0.add_worksheet('p_boiler')  
    
    worksheet0_34 = workbook0.add_worksheet('Original_Cap_cost_wind')
    worksheet0_35 = workbook0.add_worksheet('Original_Cap_cost_pv')  
    worksheet0_36 = workbook0.add_worksheet('Original_Cap_cost_wave')
    worksheet0_37 = workbook0.add_worksheet('Original_Cap_cost_turbine')
    worksheet0_38 = workbook0.add_worksheet('Original_Cap_cost_boiler')
    worksheet0_39 = workbook0.add_worksheet('Original_Cap_cost_Sbatt')
    worksheet0_40 = workbook0.add_worksheet('Original_Cap_cost_Stherm')
    worksheet0_41 = workbook0.add_worksheet('Original_Cap_cost_Swater')
    worksheet0_42 = workbook0.add_worksheet('Original_Cap_cost_Shydro')
    worksheet0_43 = workbook0.add_worksheet('Original_Cap_cost_electzer')
    worksheet0_44 = workbook0.add_worksheet('Original_Cap_cost_desal')
    
    worksheet0_45 = workbook0.add_worksheet('Mtn_cost_wind')
    worksheet0_46 = workbook0.add_worksheet('Mtn_cost_pv') 
    worksheet0_47 = workbook0.add_worksheet('Mtn_cost_wave')
    worksheet0_48 = workbook0.add_worksheet('Mtn_cost_turbine ')
    worksheet0_49 = workbook0.add_worksheet('Mtn_cost_boiler')
    worksheet0_50 = workbook0.add_worksheet('Mtn_cost_Sbatt')
    worksheet0_51 = workbook0.add_worksheet('Mtn_cost_Stherm')
    worksheet0_52 = workbook0.add_worksheet('Mtn_cost_Swater')
    worksheet0_53 = workbook0.add_worksheet('Mtn_cost_Shydro')
    worksheet0_54 = workbook0.add_worksheet('Mtn_cost_electzerr')
    worksheet0_55 = workbook0.add_worksheet('Mtn_cost_desal')
    

    
    worksheet0_56 = workbook0.add_worksheet('pWind')
    worksheet0_57 = workbook0.add_worksheet('pSolar')
    worksheet0_58 = workbook0.add_worksheet('pWave')
    
    worksheet0_59 = workbook0.add_worksheet('Annualized_Cost_invest')
    worksheet0_60 = workbook0.add_worksheet('Cost_oper')    
    worksheet0_66 = workbook0.add_worksheet('Cost_mtn') 
    
    worksheet0_61 = workbook0.add_worksheet('ds_water') 
    
    worksheet0_62 = workbook0.add_worksheet('P_Sbatt')
    worksheet0_63 = workbook0.add_worksheet('P_Stherm')
    worksheet0_64 = workbook0.add_worksheet('V_Shydro')
    worksheet0_65 = workbook0.add_worksheet('V_Swater')  
    
    # worksheet0_67 = workbook0.add_worksheet('Delta_p')
    # worksheet0_68 = workbook0.add_worksheet('Delta_h')
    # worksheet0_69 = workbook0.add_worksheet('Delta_w')
    # worksheet0_70 = workbook0.add_worksheet('Delta_q')
    
    worksheet0_71 = workbook0.add_worksheet('Annualized_Cap_cost_wind')
    worksheet0_72 = workbook0.add_worksheet('Annualized_Cap_cost_pv')  
    worksheet0_73 = workbook0.add_worksheet('Annualized_Cap_cost_wave')
    worksheet0_74 = workbook0.add_worksheet('Annualized_Cap_cost_turbine')
    worksheet0_75 = workbook0.add_worksheet('Annualized_Cap_cost_boiler')
    worksheet0_76 = workbook0.add_worksheet('Annualized_Cap_cost_Sbatt')
    worksheet0_77 = workbook0.add_worksheet('Annualized_Cap_cost_Stherm')
    worksheet0_78 = workbook0.add_worksheet('Annualized_Cap_cost_Swater')
    worksheet0_79 = workbook0.add_worksheet('Annualized_Cap_cost_Shydro')
    worksheet0_80 = workbook0.add_worksheet('Annualized_Cap_cost_electzer')
    worksheet0_81 = workbook0.add_worksheet('Annualized_Cap_cost_desal')
    
    obj_value_iter = []
    for e in range(Exp_num):
        model = gp.Model("MES")
        para = Parameter()
        
        "define variables"
        "objective"
        Original_Cost_invest = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Original_Cost_invest")
        Annualized_Cost_invest = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Annualized_Cost_invest")
        Cost_oper = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Cost_oper")
        Cost_mtn = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Cost_mtn")
        obj_value = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="obj_value")
        myAuxVar = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="myAuxVar")
        
        "variables for wind turbine"
        N_wind = model.addVar(lb=0, vtype=GRB.INTEGER, name="N_wind")
        N_wind.lb = 46
        N_wind.ub = 46
        Original_Cap_cost_wind = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Original_Cap_cost_wind")
        Annualized_Cap_cost_wind = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Annualized_Cap_cost_wind")
        Mtn_cost_wind = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Mtn_cost_wind")
        
        "variables for pv panel"
        A_pv = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="A_pv")
        A_pv.lb = 394074.2405
        A_pv.ub = 394074.2405
        Original_Cap_cost_pv = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Original_Cap_cost_pv")
        Annualized_Cap_cost_pv = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Annualized_Cap_cost_pv")
        Mtn_cost_pv = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Mtn_cost_pv")
        
        "variables for wave converter"
        N_wave = model.addVar(lb=0, vtype=GRB.INTEGER, name="N_wave")
        N_wave.lb = 0
        N_wave.ub = 0
        Original_Cap_cost_wave = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Original_Cap_cost_wave")
        Annualized_Cap_cost_wave = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Annualized_Cap_cost_wave")
        Mtn_cost_wave = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Mtn_cost_wave")

        "variables for water desalination"
        N_osmosis = model.addVar(lb=0, vtype=GRB.INTEGER, name="N_osmosis")
        N_osmosis.lb = 3
        N_osmosis.ub = 3
        cd_water = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="cd_water")
        Original_Cap_cost_desal = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Original_Cap_cost_desal")
        Annualized_Cap_cost_desal = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Annualized_Cap_cost_desal")
        water_flow_bound = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="water_flow_bound")
        Mtn_cost_desal = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Mtn_cost_desal")

        "variables for CHP system"
        p_turbine = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="p_turbine")
        p_turbine.lb = 22.06639188
        p_turbine.ub = 22.06639188
        p_boiler = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="p_boiler")
        p_boiler.lb = 18.35844645
        p_boiler.ub = 18.35844645
        Original_Cap_cost_turbine = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Original_Cap_cost_turbine")
        Annualized_Cap_cost_turbine = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Annualized_Cap_cost_turbine")
        Mtn_cost_turbine = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Mtn_cost_turbine")
        Original_Cap_cost_boiler = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Original_Cap_cost_boiler")
        Annualized_Cap_cost_boiler = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Annualized_Cap_cost_boiler")
        Mtn_cost_boiler = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Mtn_cost_boiler")
        
        "variables for electrolyzer"
        Original_Cap_cost_electzer = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Original_Cap_cost_electzer")
        Annualized_Cap_cost_electzer = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Annualized_Cap_cost_electzer")
        N_electzer = model.addVar(lb=0, vtype=GRB.INTEGER, name="N_electzer")
        N_electzer.lb = 5
        N_electzer.ub = 5
        Mtn_cost_electzer = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Mtn_cost_electzer")

        # Variables for storages and tanks
        p_Sbatt = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="p_Sbatt")
        p_Stherm = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="p_Stherm")
        v_Swater = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="v_Swater")
        v_Shydro = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="v_Shydro")

        p_Sbatt.lb = 209.7210511
        p_Sbatt.ub = 209.7210511
        p_Stherm.lb = 0
        p_Stherm.ub = 0
        v_Swater.lb = 37225.8625
        v_Swater.ub = 37225.8625
        v_Shydro.lb = 6450.639867
        v_Shydro.ub = 6450.639867



        Original_Cap_cost_Sbatt = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Original_Cap_cost_Sbatt")
        Annualized_Cap_cost_Sbatt = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Annualized_Cap_cost_Sbatt")
        Original_Cap_cost_Stherm = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Original_Cap_cost_Stherm")
        Annualized_Cap_cost_Stherm = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Annualized_Cap_cost_Stherm")
        Original_Cap_cost_Swater = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Original_Cap_cost_Swater")
        Annualized_Cap_cost_Swater = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Annualized_Cap_cost_Swater")
        Original_Cap_cost_Shydro = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Original_Cap_cost_Shydro")
        Annualized_Cap_cost_Shydro = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Annualized_Cap_cost_Shydro")
        Mtn_cost_Sbatt = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Mtn_cost_Sbatt")
        Mtn_cost_Stherm = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Mtn_cost_Stherm")
        Mtn_cost_Swater = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Mtn_cost_Swater")
        Mtn_cost_Shydro = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name="Mtn_cost_Shydro") 
        
        b_x, q_x, w_x, h_x = {}, {}, {}, {}
        for t in range(para.Time_period):
            b_x[t] = model.addVar(lb=-GRB.INFINITY, vtype=GRB.CONTINUOUS, name=f"b_x_{t}")  
            q_x[t] = model.addVar(lb=-GRB.INFINITY, vtype=GRB.CONTINUOUS, name=f"q_x_{t}") 
            w_x[t] = model.addVar(lb=-GRB.INFINITY, vtype=GRB.CONTINUOUS, name=f"w_x_{t}") 
            h_x[t] = model.addVar(lb=-GRB.INFINITY, vtype=GRB.CONTINUOUS, name=f"h_x_{t}") 

            
        
        rPow, wFlow, eP_des, q1GT, q2GT, pGT, gFlowP, gFlowB, hFlow, eP_eltz = {}, {}, {}, {}, {}, {}, {}, {}, {}, {}
        pWind, pSolar, dV, pWave, ds_water = {},{},{},{},{}
        
        for t in range(para.Time_period):
            rPow[t] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"rPow_{t}")
            
            # Variables for wind turbine
            pWind[t] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"pWind_{t}")
            
            # Variables for solar power
            pSolar[t] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"pSolar_{t}")
            
            # Variables for wave converter
            dV[t] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"dV_{t}")
            pWave[t] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"pWave_{t}")
            
            # Variables for water desalination
            wFlow[t] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"wFlow_{t}")
            eP_des[t] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"eP_des_{t}")
            ds_water[t] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"ds_water_{t}")

            # Variables for CHP system
            pGT[t] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"pGT_{t}")
            q1GT[t] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"q1GT_{t}")
            q2GT[t] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"q2GT_{t}")
            gFlowP[t] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"gFlowP_{t}")
            gFlowB[t] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"gFlowB_{t}")
            
            # Variables for electrolyzer
            hFlow[t] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"hFlow_{t}")
            eP_eltz[t] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"eP_eltz_{t}")
            
        b_xd, b_xc, q_xd, q_xc, w_xd, w_xc, h_xd, h_xc =   {}, {}, {}, {}, {}, {}, {}, {}
        sB,sQ,sW,sH = {}, {}, {}, {}
        for t in range(para.Time_period):
            b_xd[t] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"b_xd_{t}")  # 'power discharged by battery'
            b_xc[t] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"b_xc_{t}")  # 'power charged to battery'
            q_xd[t] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"q_xd_{t}")  # 'thermal discharged by thermal storage'
            q_xc[t] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"q_xc_{t}")  # 'thermal charged to thermal storage'
            w_xd[t] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"w_xd_{t}")  # 'water discharged by water tank'
            w_xc[t] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"w_xc_{t}")  # 'water charged to water tank'
            h_xd[t] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"h_xd_{t}")  # 'hydrogen discharged by hydrogen tank'
            h_xc[t] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"h_xc_{t}")  # 'hydrogen charged to hydrogen tank'
            sB[t] = model.addVar(lb=0, ub=1, vtype=GRB.CONTINUOUS, name=f"sB_{t}")  # 'state of charge, battery'
            sQ[t] = model.addVar(lb=0, ub=1, vtype=GRB.CONTINUOUS, name=f"sQ_{t}")  # 'state of charge, thermal storage'
            sW[t] = model.addVar(lb=0, ub=1, vtype=GRB.CONTINUOUS, name=f"sW_{t}")  # 'state of charge, water tank'
            sH[t] = model.addVar(lb=0, ub=1, vtype=GRB.CONTINUOUS, name=f"sH_{t}")  # 'state of charge, hydrogen tank'
        
        # Delta_p, Delta_q, Delta_h, Delta_w = {}, {}, {},{}
        # for t in range(para.Time_period):
        #     Delta_h[t] = model.addVar(lb=0, vtype="C", name="Delta_h")
        #     Delta_w[t] = model.addVar(lb=0, vtype="C", name="Delta_w")
        #     Delta_q[t] = model.addVar(lb=0, vtype="C", name="Delta_q")
        #     Delta_p[t] = model.addVar(lb=0, vtype="C", name="Delta_p")
        
        '============================================= constraints list========================================'
        'define objective '
        
        model.addConstr(Original_Cost_invest == Original_Cap_cost_wind + Original_Cap_cost_pv + Original_Cap_cost_wave + Original_Cap_cost_desal + Original_Cap_cost_turbine + Original_Cap_cost_boiler + Original_Cap_cost_electzer + 
                    Original_Cap_cost_Sbatt + Original_Cap_cost_Stherm + Original_Cap_cost_Swater + Original_Cap_cost_Shydro)
        
        model.addConstr(Annualized_Cost_invest == Original_Cost_invest * (para.interest_rate * (1 + para.interest_rate) ** para.life_time)/((1 + para.interest_rate) ** para.life_time - 1))
        
        model.addConstr(Cost_oper == gp.quicksum((gFlowP[t] + gFlowB[t]) * para.gas_price for t in range(para.Time_period)) +
                    gp.quicksum((gFlowP[t] + gFlowB[t]) * para.carbon_price * para.gas_emmi_factor for t in range(para.Time_period))) 
                    #quicksum(para.Lambda_H * Delta_h[t] + para.Lambda_Q * Delta_q[t] + para.Lambda_W * Delta_w[t] + para.Elec_price[t] * Delta_p[t]  for t in range(para.Time_period)))
    #                  quicksum(para.gamma_B * abs(b_x[t]) + para.gamma_Q * abs(q_x[t]) + para.gamma_H * abs(h_x[t]) + para.gamma_W * abs(w_x[t]) for t in range(para.Time_period)))

        model.addConstr(Cost_mtn == Mtn_cost_wind + Mtn_cost_pv + Mtn_cost_wave + Mtn_cost_desal + Mtn_cost_turbine + Mtn_cost_boiler + Mtn_cost_electzer + 
                    Mtn_cost_Sbatt + Mtn_cost_Stherm + Mtn_cost_Swater + Mtn_cost_Shydro)
        
        
        model.addConstr(obj_value == Annualized_Cost_invest + (Cost_oper * 365/(para.Time_period/24)) + Cost_mtn)
        
        
        #************************************************************ renewable power *********************************************** 
        for t in range(para.Time_period):
            model.addConstr(rPow[t] == pWind[t] + pSolar[t] + pWave[t]) 
            
        # model.addConstr(gp.quicksum(rPow[t] for t in range(para.Time_period)) >=0.9*(gp.quicksum(rPow[t] for t in range(para.Time_period)) + gp.quicksum(pGT[t] for t in range(para.Time_period))))
        # model.addCons(quicksum(rPow[t] for t in range(para.Time_period))/(quicksum(rPow[t] for t in range(para.Time_period)) + quicksum(pGT[t] for t in range(para.Time_period))) <= 1)
        
        'define the constraints for wind turbine'
        for t in range(para.Time_period):
            if (para.VWind[t] < para.I_w) or (para.VWind[t] >= para.Phi_w):
                model.addConstr(pWind[t] == 0)
            elif (para.VWind[t] >= para.I_w) and (para.VWind[t] < para.Lambda_w):
                model.addConstr(pWind[t] == N_wind * para.Rated_w * ((para.VWind[t] - para.I_w)/ (para.Lambda_w - para.I_w)))
            elif (para.VWind[t] >= para.Lambda_w) and (para.VWind[t] <= para.Phi_w):
                model.addConstr(pWind[t] == N_wind * para.Rated_w)
                    
        model.addConstr(Original_Cap_cost_wind == N_wind * para.Rated_w * para.Ucost_w)
        model.addConstr(Annualized_Cap_cost_wind == Original_Cap_cost_wind * (para.interest_rate * (1 + para.interest_rate) ** para.life_time)/((1 + para.interest_rate) ** para.life_time - 1))
        model.addConstr(Mtn_cost_wind == N_wind * para.Rated_w * para.Mcost_w)
        
        'define the constraints for pv panel'
        for t in range(para.Time_period):
            model.addConstr(pSolar[t] == para.solar_irri[t] * para.tal_PV * para.eta_PV * A_pv * (1 - para.gamma_PV * (para.temp[t] - 25)))
            # model.addConstr(A_pv <= 4000)
        model.addConstr(Original_Cap_cost_pv == A_pv * para.Ucost_PV) 
        model.addConstr(Annualized_Cap_cost_pv == Original_Cap_cost_pv * (para.interest_rate * (1 + para.interest_rate) ** para.life_time)/((1 + para.interest_rate) ** para.life_time - 1))
        model.addConstr(Mtn_cost_pv == para.cap_pv_m2 * A_pv * para.Mcost_PV)
        
        'define the constraints for wave converter'
        for t in range(para.Time_period):
            model.addConstr(pWave[t] == N_wave * para.eta_wave * 0.000001 * para.water_density * para.gravity_accel * para.gravity_accel * para.wave_time * para.HWave[t] * para.HWave[t] * para.water_front/(32 * math.pi))
            # model.addCons(dV[t] == para.water_density * para.gravity_accel * para.HWave[t] * para.HWave[t] * math.sqrt(para.gravity_accel * para.water_depth)/16)
            # model.addCons(pWave[t] == dV[t] * para.eta_wave * para.length_wave * N_wave/1000) # dv is kw/m, convert to mw
        
        model.addConstr(Original_Cap_cost_wave == N_wave * para.Ucost_wave)
        model.addConstr(Annualized_Cap_cost_wave == Original_Cap_cost_wave * (para.interest_rate * (1 + para.interest_rate) ** para.life_time)/((1 + para.interest_rate) ** para.life_time - 1))
        model.addConstr(Mtn_cost_wave == Original_Cap_cost_wave * para.Mcost_wave)
        
        #************************************************************ desalination *************************************************   
        for t in range(para.Time_period):
            model.addConstr(eP_des[t] * 1000 == ds_water[t] * wFlow[t])
            model.addConstr(ds_water[t] == 2.05 * 0.00001 * para.C0 * ((2-para.eta_des)/(2*(1-para.eta_des))) + 2.78 * 0.0000001 * para.mem_resist * wFlow[t])
            model.addConstr(wFlow[t] <= para.Water_flow_bound * N_osmosis)
            model.addConstr(cd_water == 4472.94 * (24 * para.Water_flow_bound)**(-0.125))
        
        model.addConstr(Original_Cap_cost_desal == cd_water * para.Water_flow_bound * 24 * N_osmosis)
        model.addConstr(Annualized_Cap_cost_desal == Original_Cap_cost_desal *(para.interest_rate * (1 + para.interest_rate) ** para.life_time)/((1 + para.interest_rate) ** para.life_time - 1))
        model.addConstr(Mtn_cost_desal == Original_Cap_cost_desal * para.Mcost_des)
    
        #************************************************************ gas turbine ************************************************* 
        for t in range(para.Time_period):
            model.addConstr(pGT[t] * para.DeltaT == gFlowP[t] * para.Vgas * para.eta_P)
            model.addConstr(q1GT[t] <= para.eta_X * pGT[t] * (1 - para.eta_P - para.eta_L)/para.eta_P)
            model.addConstr(q2GT[t] * para.DeltaT <= gFlowB[t] * para.Vgas * para.eta_Q)
            model.addConstr(pGT[t] <= p_turbine)
            model.addConstr(q2GT[t] <= p_boiler)
        
        model.addGenConstrPow(p_turbine, myAuxVar, 0.57, "gf", "FuncPieces=1000")
        model.addConstr(Original_Cap_cost_turbine == 13885.57 * (myAuxVar))  #On the value of combined heat and power (CHP) systems and heat pumps in centralised and distributed heating systems: Lessons from multi-fidelity modelling approaches
        model.addConstr(Annualized_Cap_cost_turbine == Original_Cap_cost_turbine *(para.interest_rate * (1 + para.interest_rate) ** para.life_time)/((1 + para.interest_rate) ** para.life_time - 1))
        # model.addCons(Cap_cost_boiler == 221.7124 * (p_boiler ** 0.707))
        # model.addCons(Cap_cost_turbine == (-138.71* log(p_turbine) + 1727.1) * 1000) # Modeling and optimizing a CHP system for natural gas pressure reduction plant
        # model.addCons(Cap_cost_turbine == 28641 * (p_turbine ** (-0.2454)) * 1000 * 0.145) # $/MW, see paper "Study on the optimization and sensitivity analysis of CCHP systems for T industrial park facilities"
        model.addConstr(Original_Cap_cost_boiler == 200 * 1000 * 0.145 * p_boiler) # $/MW see table 2-3 see paper "Study on the optimization and sensitivity analysis of CCHP systems for T industrial park facilities"
        model.addConstr(Annualized_Cap_cost_boiler == Original_Cap_cost_boiler *(para.interest_rate * (1 + para.interest_rate) ** para.life_time)/((1 + para.interest_rate) ** para.life_time - 1))
        
        model.addConstr(Mtn_cost_turbine == Original_Cap_cost_turbine * para.Mcost_turbine)   # see table 2-3 see paper "Study on the optimization and sensitivity analysis of CCHP systems for T industrial park facilities"
        model.addConstr(Mtn_cost_boiler == p_boiler * para.Mcost_boiler) # see table 2-3 see paper "Study on the optimization and sensitivity analysis of CCHP systems for T industrial park facilities"
        
        
        #************************************************************ electrolyzer ************************************************* 
        for t in range(para.Time_period):   
            model.addConstr(hFlow[t] == para.eta_eltz * eP_eltz[t] * para.PH * N_electzer)
            model.addConstr(para.alpha_lb * para.rated_electzer <= eP_eltz[t])
            model.addConstr(eP_eltz[t] <= para.rated_electzer)
        model.addConstr(Original_Cap_cost_electzer == N_electzer * para.rated_electzer * para.ce_electzer)
        model.addConstr(Annualized_Cap_cost_electzer == Original_Cap_cost_electzer *(para.interest_rate * (1 + para.interest_rate) ** para.life_time)/((1 + para.interest_rate) ** para.life_time - 1))
        
        model.addConstr(Mtn_cost_electzer == Original_Cap_cost_electzer * para.Mcost_electzer)     
            
        #************************************************************ battery *************************************************  
        for t in range(para.Time_period): 
            model.addConstr(b_x[t] == b_xd[t] - b_xc[t])    
        
        model.addConstr(sB[0] * p_Sbatt == para.IB * p_Sbatt - ((b_xd[0]/para.eta_bd)) * para.DeltaT + ((b_xc[0] * para.eta_bc)) * para.DeltaT)

        for t in range(1,para.Time_period):
            model.addConstr(sB[t] * p_Sbatt == sB[t-1] * p_Sbatt - ((b_xd[t]/para.eta_bd)) * para.DeltaT + ((b_xc[t] * para.eta_bc)) * para.DeltaT)
            # if t == para.Time_period - 1:
            #     model.addCons(sB[t] >= 0.2)
            if t in para.end_day_time_list:
                model.addConstr(sB[t] >= para.IB)
            
        for t in range(para.Time_period):
            model.addConstr(b_x[t] <= p_Sbatt * para.cdr)
            model.addConstr(b_x[t] >= -p_Sbatt * para.cdr)

        
        model.addConstr(Original_Cap_cost_Sbatt == p_Sbatt * para.Ucost_Sbatt)
        model.addConstr(Annualized_Cap_cost_Sbatt == Original_Cap_cost_Sbatt *(para.interest_rate * (1 + para.interest_rate) ** para.life_time)/((1 + para.interest_rate) ** para.life_time - 1))
        model.addConstr(Mtn_cost_Sbatt == p_Sbatt * para.Mcost_Sbatt) 
        
        # ************************************************************ thermal storage *********************************************  
        for t in range(para.Time_period): 
            model.addConstr(q_x[t] == q_xd[t] - q_xc[t])    
        
        model.addConstr(sQ[0] * p_Stherm == para.IQ * p_Stherm - ((q_xd[0]/para.eta_qd)) * para.DeltaT + ((q_xc[0] * para.eta_qc)) * para.DeltaT)

        for t in range(1,para.Time_period):
            model.addConstr(sQ[t] * p_Stherm == sQ[t-1] * p_Stherm - ((q_xd[t]/para.eta_qd)) * para.DeltaT + ((q_xc[t] * para.eta_qc)) * para.DeltaT)
            # if t == para.Time_period - 1:
            #     model.addCons(sQ[t] >= 0.2)
            if t in para.end_day_time_list:
                model.addConstr(sQ[t] >= para.IQ)
                
        for t in range(para.Time_period):
            model.addConstr(q_x[t] <= p_Stherm * para.cdr)
            model.addConstr(q_x[t] >= -p_Stherm * para.cdr) 
        
        model.addConstr(Original_Cap_cost_Stherm == p_Stherm * para.Ucost_Stherm)
        model.addConstr(Annualized_Cap_cost_Stherm == Original_Cap_cost_Stherm *(para.interest_rate * (1 + para.interest_rate) ** para.life_time)/((1 + para.interest_rate) ** para.life_time - 1))
        model.addConstr(Mtn_cost_Stherm == Original_Cap_cost_Stherm * para.Mcost_Stherm)
        
        #************************************************************ water tank *************************************************  
        for t in range(para.Time_period): 
            model.addConstr(w_x[t] == w_xd[t] - w_xc[t])    
        
        model.addConstr(sW[0] * v_Swater == para.IW * v_Swater - ((w_xd[0]/para.eta_wd)) * para.DeltaT + ((w_xc[0] * para.eta_wc)) * para.DeltaT)

        for t in range(1,para.Time_period):
            model.addConstr(sW[t] * v_Swater == sW[t-1] * v_Swater - ((w_xd[t]/para.eta_wd)) * para.DeltaT + ((w_xc[t] * para.eta_wc)) * para.DeltaT)
            # model.addConstr(v_Swater <= 1000) #put limitation on water storage https://www2.deq.idaho.gov/admin/LEIA/api/document/download/4791 (5.4) x Qmxdy/24 #https://www.snyder-associates.com/community-water-storage-solutions-how-to-select-a-water-tank/
            
            # if t == para.Time_period - 1:
            #     model.addCons(sW[t] >= 0.2)
            #initial charge
            if t in para.end_day_time_list:
                model.addConstr(sW[t] >= para.IW)
        
        for t in range(para.Time_period):
            model.addConstr(w_x[t] <= v_Swater * para.cdr)
            model.addConstr(w_x[t] >= -v_Swater * para.cdr) 
            
        model.addConstr(Original_Cap_cost_Swater == v_Swater * para.Ucost_Swater) 
        model.addConstr(Annualized_Cap_cost_Swater == Original_Cap_cost_Swater * (para.interest_rate * (1 + para.interest_rate) ** para.life_time)/((1 + para.interest_rate) ** para.life_time - 1)) 
        model.addConstr(Mtn_cost_Swater == Original_Cap_cost_Swater * para.Mcost_Swater)

        #************************************************************ hydrogen tank ***********************************************   
        for t in range(para.Time_period): 
            model.addConstr(h_x[t] == h_xd[t] - h_xc[t])    
        
        model.addConstr(sH[0] * v_Shydro == para.IH * v_Shydro - ((h_xd[0]/para.eta_hd)) * para.DeltaT + ((h_xc[0] * para.eta_hc)) * para.DeltaT)

        for t in range(1,para.Time_period):
            model.addConstr(sH[t] * v_Shydro == sH[t-1] * v_Shydro - ((h_xd[t]/para.eta_hd)) * para.DeltaT + ((h_xc[t] * para.eta_hc)) * para.DeltaT)
            # model.addConstr(v_Shydro <= 5000) #put limitation on hydrogen storage Optimizing green hydrogen systems: Balancing economic viability and reliability in the face of supply-demand volatility
            
            # if t == para.Time_period - 1:
            #     model.addCons(sH[t] >= 0.2)
            if t in para.end_day_time_list:
                model.addConstr(sH[t] >= para.IH)
        
        for t in range(para.Time_period):
            model.addConstr(h_x[t] <= v_Shydro * para.cdr)
            model.addConstr(h_x[t] >= -v_Shydro * para.cdr) 
        
        model.addConstr(Original_Cap_cost_Shydro == v_Shydro * para.Ucost_Shydro) 
        model.addConstr(Annualized_Cap_cost_Shydro == Original_Cap_cost_Shydro * (para.interest_rate * (1 + para.interest_rate) ** para.life_time)/((1 + para.interest_rate) ** para.life_time - 1)) 
        model.addConstr(Mtn_cost_Shydro ==  v_Shydro * para.Mcost_Shydro)

        #************************************************************ energy balance **********************************************   
        for t in range(para.Time_period):
            model.addConstr(wFlow[t] + w_x[t]  - para.wD[t] == 0)
            
        for t in range(para.Time_period):
            model.addConstr(q1GT[t] + q2GT[t] + q_x[t]  - para.qD[t] == 0)

        for t in range(para.Time_period):
            model.addConstr(hFlow[t] + h_x[t]  - para.hD[t] == 0)
        
        for t in range(para.Time_period):
            model.addConstr(rPow[t] + pGT[t] + b_x[t]  - eP_des[t] - eP_eltz[t] - para.eD[t] == 0)

    
        '============================================= solving process ========================================'
        model.setObjective(obj_value, GRB.MINIMIZE)
        # Set the MIP gap tolerance
        model.Params.MIPGap = 0.001
        # Display output
        model.Params.OutputFlag = 1
        
        try:
            model.optimize()
            end_time = time.time()
            
            print("=========================")
            print(obj_value.X)
            obj_value_iter.append(obj_value.X)
        except:
            pass

            # print('time needed', end_time-begin_time)
            
                # workbook0 = xlsxwriter.Book('MES_result_capacity.xlsx')
                # s=workbook0.add_sheet()
                # for t in range(para.Time_period):
                #     s.write(t)
                # workbook0.close()     

    #         for t in range(para.Time_period):
    #             worksheet0_0.write(e, 0, obj_value.X)
    #             worksheet0_1.write(e, t, b_x[t].X)
    #             worksheet0_2.write(e, t, q_x[t].X)
    #             worksheet0_3.write(e, t, w_x[t].X)
    #             worksheet0_4.write(e, t, h_x[t].X)
    #             worksheet0_5.write(e, t, b_xd[t].X)
    #             worksheet0_6.write(e, t, b_xc[t].X)
    #             worksheet0_7.write(e, t, q_xd[t].X)
    #             worksheet0_8.write(e, t, q_xc[t].X)
    #             worksheet0_9.write(e, t, w_xd[t].X)
    #             worksheet0_10.write(e, t, w_xc[t].X)
    #             worksheet0_11.write(e, t, h_xd[t].X)
    #             worksheet0_12.write(e, t, h_xc[t].X)   
    #             worksheet0_13.write(e, t, rPow[t].X)
    #             worksheet0_14.write(e, t, wFlow[t].X)
    #             worksheet0_15.write(e, t, eP_des[t].X)
    #             worksheet0_16.write(e, t, q1GT[t].X)
    #             worksheet0_17.write(e, t, q2GT[t].X)
    #             worksheet0_18.write(e, t, pGT[t].X)
    #             worksheet0_19.write(e, t, gFlowP[t].X)
    #             worksheet0_20.write(e, t, gFlowB[t].X)
    #             worksheet0_21.write(e, t, hFlow[t].X)
    #             worksheet0_22.write(e, t, eP_eltz[t].X)   
    #             worksheet0_23.write(e, t, sB[t].X)
    #             worksheet0_24.write(e, t, sQ[t].X)
    #             worksheet0_25.write(e, t, sW[t].X)
    #             worksheet0_26.write(e, t, sH[t].X) 

    #             worksheet0_56.write(e, t, pWind[t].X)
    #             worksheet0_57.write(e, t, pSolar[t].X)
    #             worksheet0_58.write(e, t, pWave[t].X) 
    #             worksheet0_61.write(e, t, ds_water[t].X)

                
    #             # worksheet0_67.write(0,t, Delta_p[t]))
    #             # worksheet0_68.write(0,t, Delta_h[t]))
    #             # worksheet0_69.write(0,t, Delta_w[t]))     
    #             # worksheet0_70.write(0,t, Delta_q[t])) 


                
                
    #         worksheet0_27.write(e,0, N_wind.X) 
    #         worksheet0_28.write(e,0, A_pv.X) 
    #         worksheet0_29.write(e,0, N_wave.X) 
    #         worksheet0_30.write(e,0, N_osmosis.X) 
    #         worksheet0_31.write(e,0, N_electzer.X) 
    #         worksheet0_32.write(e,0, p_turbine.X) 
    #         worksheet0_33.write(e,0, p_boiler.X)  

            
    #         worksheet0_34.write(e,0, Original_Cap_cost_wind.X) 
    #         worksheet0_35.write(e,0, Original_Cap_cost_pv.X) 
    #         worksheet0_36.write(e,0, Original_Cap_cost_wave.X) 
    #         worksheet0_37.write(e,0, Original_Cap_cost_turbine.X) 
    #         worksheet0_38.write(e,0, Original_Cap_cost_boiler.X) 
    #         worksheet0_39.write(e,0, Original_Cap_cost_Sbatt.X) 
    #         worksheet0_40.write(e,0, Original_Cap_cost_Stherm.X)  
    #         worksheet0_41.write(e,0, Original_Cap_cost_Swater.X) 
    #         worksheet0_42.write(e,0, Original_Cap_cost_Shydro.X)       
    #         worksheet0_43.write(e,0, Original_Cap_cost_electzer.X)   
    #         worksheet0_44.write(e,0, Original_Cap_cost_desal.X) 
        
    #         worksheet0_45.write(e,0, Mtn_cost_wind.X) 
    #         worksheet0_46.write(e,0, Mtn_cost_pv.X) 
    #         worksheet0_47.write(e,0, Mtn_cost_wave.X) 
    #         worksheet0_48.write(e,0, Mtn_cost_turbine.X) 
    #         worksheet0_49.write(e,0, Mtn_cost_boiler.X) 
    #         worksheet0_50.write(e,0, Mtn_cost_Sbatt.X) 
    #         worksheet0_51.write(e,0, Mtn_cost_Stherm.X)  
    #         worksheet0_52.write(e,0, Mtn_cost_Swater.X) 
    #         worksheet0_53.write(e,0, Mtn_cost_Shydro.X)       
    #         worksheet0_54.write(e,0, Mtn_cost_electzer.X)   
    #         worksheet0_55.write(e,0, Mtn_cost_desal.X)
            
    #         worksheet0_59.write(e,0, Annualized_Cost_invest.X)   
    #         worksheet0_60.write(e,0, Cost_oper.X) 
    #         worksheet0_66.write(e,0, Cost_mtn.X) 
            
    #         worksheet0_62.write(e,0,p_Sbatt.X)
    #         worksheet0_63.write(e,0,p_Stherm.X)
    #         worksheet0_64.write(e,0,v_Shydro.X)     
    #         worksheet0_65.write(e,0,v_Swater.X)      

    #         worksheet0_71.write(e,0, Annualized_Cap_cost_wind.X) 
    #         worksheet0_72.write(e,0, Annualized_Cap_cost_pv.X) 
    #         worksheet0_73.write(e,0, Annualized_Cap_cost_wave.X) 
    #         worksheet0_74.write(e,0, Annualized_Cap_cost_turbine.X) 
    #         worksheet0_75.write(e,0, Annualized_Cap_cost_boiler.X) 
    #         worksheet0_76.write(e,0, Annualized_Cap_cost_Sbatt.X) 
    #         worksheet0_77.write(e,0, Annualized_Cap_cost_Stherm.X)  
    #         worksheet0_78.write(e,0, Annualized_Cap_cost_Swater.X) 
    #         worksheet0_79.write(e,0, Annualized_Cap_cost_Shydro.X)       
    #         worksheet0_80.write(e,0, Annualized_Cap_cost_electzer.X)   
    #         worksheet0_81.write(e,0, Annualized_Cap_cost_desal.X)   

    #     except:
    #         pass
    # workbook0.close()   
    print("**********************")
    print(obj_value_iter)




    







    
        